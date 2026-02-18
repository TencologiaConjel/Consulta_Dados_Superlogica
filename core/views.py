from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.views import View
from django.db.models import Q, Sum
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from datetime import datetime
from decimal import Decimal, InvalidOperation
import re
import openpyxl

from .models import Receitas, Despesas


# View Login User
def login_usuario(request):
    if request.method == "POST":
        email = request.POST.get("email")
        senha = request.POST.get("senha")

        if not email or not senha:
            messages.error(request, "Por favor, preencha todos os campos.")
            return render(request, "login.html")

        User = get_user_model()
        try:
            User.objects.get(email=email)
            usuario = authenticate(request, email=email, password=senha)
        except User.DoesNotExist:
            usuario = None

        if usuario is not None:
            login(request, usuario)
            messages.success(request, "Login realizado com sucesso!")
            return redirect("dashboard")

        messages.error(request, "Email ou senha inválidos.")

    return render(request, "login.html")


def to_decimal(v):
    if v is None:
        return Decimal("0.00")

    if isinstance(v, Decimal):
        return v

    if isinstance(v, (int, float)):
        return Decimal(str(v))

    s = str(v).strip()
    if s == "" or s.lower() in {"-", "—", "null", "none", "nan"}:
        return Decimal("0.00")

    s = s.replace("R$", "").replace("\u00a0", " ").strip()

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = re.sub(r"[^0-9,.\-]", "", s)

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    else:
        if "," in s:
            s = s.replace(".", "").replace(",", ".")

    try:
        val = Decimal(s)
        return -val if neg else val
    except (InvalidOperation, ValueError):
        raise ValueError(f"Valor inválido para Decimal: {v} (normalizado: {s})")


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%Y"):
        try:
            d = datetime.strptime(s, fmt)
            return d.date()
        except ValueError:
            pass
    raise ValueError(f"Data inválida: {v}")


def is_row_empty(row):
    for v in row:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    s = s.replace("ú", "u")
    s = s.replace("ç", "c")
    s = re.sub(r"\s+", "", s)
    return s


def detect_condominio(texto):
    """
    Aceita SOMENTE o padrão do relatório:
    '078A Basualdo (6)'  -> retorna 'Basualdo'
    Qualquer outra coisa retorna None (evita capturar datas e textos da tabela)
    """
    if texto is None:
        return None

    # nunca tratar datas como condomínio
    if isinstance(texto, (datetime, )):
        return None
    if hasattr(texto, "year") and hasattr(texto, "month") and hasattr(texto, "day"):
        return None

    s = str(texto).strip()
    if not s:
        return None

    # precisa ter "(numero)" no final, senão não é cabeçalho de condomínio
    if "(" not in s or ")" not in s:
        return None

    s = re.sub(r"\s+", " ", s).strip()

    m = re.match(r"^[0-9A-Za-z]+\s+(.+?)\s*\(\s*\d+\s*\)\s*$", s)
    if not m:
        return None

    return m.group(1).strip()



class DashboardView(View):
    template_name = "dashboard.html"

    def get(self, request):
        tab = (request.GET.get("tab") or "receitas").strip().lower()
        if tab not in {"receitas", "despesas"}:
            tab = "receitas"

        context = {"tab": tab}

        if tab == "receitas":
            qs = (
                Receitas.objects
                .all()
                .only(
                    "NomeCondominio",
                    "Unidade",
                    "Contato",
                    "DescricaoTaxa",
                    "Cobranca",
                    "Valor",
                    "Vencimento",
                    "Liquidacao",
                )
                .order_by("Cobranca", "-Vencimento")
            )

            cliente = (request.GET.get("cliente") or "").strip()
            if cliente:
                qs = qs.filter(Q(Contato__icontains=cliente) | Q(NomeCondominio__icontains=cliente))

            v_ini = (request.GET.get("data_vencimento_inicial") or "").strip()
            if v_ini:
                d = parse_date(v_ini)
                if d:
                    qs = qs.filter(Vencimento__gte=d)

            v_fim = (request.GET.get("data_vencimento_final") or "").strip()
            if v_fim:
                d = parse_date(v_fim)
                if d:
                    qs = qs.filter(Vencimento__lte=d)

            l_ini = (request.GET.get("data_liquidacao_inicial") or "").strip()
            if l_ini:
                d = parse_date(l_ini)
                if d:
                    qs = qs.filter(Liquidacao__gte=d, Liquidacao__isnull=False)

            l_fim = (request.GET.get("data_liquidacao_final") or "").strip()
            if l_fim:
                d = parse_date(l_fim)
                if d:
                    qs = qs.filter(Liquidacao__lte=d, Liquidacao__isnull=False)

            total_receitas = qs.count()
            valor_total = qs.aggregate(total=Sum("Valor"))["total"] or Decimal("0.00")
            valor_total_formatado = f"{valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            page_number = request.GET.get("page", 1)
            paginator = Paginator(qs, 100)
            page_obj = paginator.get_page(page_number)
            receitas = list(page_obj.object_list)

            linhas = []
            cobranca_atual = None
            subtotal = Decimal("0.00")

            def norm_cobranca(v):
                return v if v is not None else 0

            def push_subtotal(cobranca, subtotal_val, add_spacer=True):
                linhas.append({
                    "tipo": "subtotal",
                    "cobranca": cobranca,
                    "subtotal": subtotal_val,
                })
                if add_spacer:
                    linhas.append({"tipo": "spacer"})

            for r in receitas:
                c = norm_cobranca(r.Cobranca)

                if cobranca_atual is None:
                    cobranca_atual = c

                if c != cobranca_atual:
                    push_subtotal(cobranca_atual, subtotal, add_spacer=True)
                    cobranca_atual = c
                    subtotal = Decimal("0.00")

                linhas.append({"tipo": "item", "obj": r})
                subtotal += (r.Valor or Decimal("0.00"))

            if cobranca_atual is not None:
                push_subtotal(cobranca_atual, subtotal, add_spacer=False)

            params = request.GET.copy()
            params.pop("page", None)
            base_qs = params.urlencode()

            context.update({
                "receitas": receitas,
                "linhas": linhas,
                "page_obj": page_obj,
                "base_qs": base_qs,
                "total_receitas": total_receitas,
                "valor_total": valor_total_formatado,
                "receitas_pagas": qs.filter(Liquidacao__isnull=False).count(),
                "receitas_pendentes": qs.filter(Liquidacao__isnull=True).count(),
            })

        else:
            qs = (
                Despesas.objects
                .all()
                .only(
                    "Condominio",
                    "Vencimento",
                    "Liquidacao",
                    "Fornecedor",
                    "Categoria",
                    "Complemento",
                    "Documento",
                    "Competencia",
                    "FormaPagamento",
                    "Valor",
                )
                .order_by("Condominio", "-Vencimento")
            )

            condominio = (request.GET.get("condominio") or "").strip()
            if condominio:
                qs = qs.filter(Condominio__icontains=condominio)

            fornecedor = (request.GET.get("fornecedor") or "").strip()
            if fornecedor:
                qs = qs.filter(Fornecedor__icontains=fornecedor)

            categoria = (request.GET.get("categoria") or "").strip()
            if categoria:
                qs = qs.filter(Categoria__icontains=categoria)

            v_ini = (request.GET.get("data_vencimento_inicial") or "").strip()
            if v_ini:
                d = parse_date(v_ini)
                if d:
                    qs = qs.filter(Vencimento__gte=d)

            v_fim = (request.GET.get("data_vencimento_final") or "").strip()
            if v_fim:
                d = parse_date(v_fim)
                if d:
                    qs = qs.filter(Vencimento__lte=d)

            total_despesas = qs.count()
            valor_total = qs.aggregate(total=Sum("Valor"))["total"] or Decimal("0.00")
            valor_total_formatado = f"{valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

            page_number = request.GET.get("page", 1)
            paginator = Paginator(qs, 100)
            page_obj = paginator.get_page(page_number)
            despesas = list(page_obj.object_list)

            linhas = []
            condominio_atual = None
            subtotal = Decimal("0.00")

            def push_subtotal_desp(cond, subtotal_val):
                linhas.append({
                    "tipo": "subtotal_despesa",
                    "condominio": cond,
                    "subtotal": subtotal_val,
                })
                linhas.append({"tipo": "spacer"})

            for d in despesas:
                if condominio_atual is None:
                    condominio_atual = d.Condominio

                if d.Condominio != condominio_atual:
                    push_subtotal_desp(condominio_atual, subtotal)
                    condominio_atual = d.Condominio
                    subtotal = Decimal("0.00")

                linhas.append({"tipo": "item_despesa", "obj": d})
                subtotal += (d.Valor or Decimal("0.00"))

            if condominio_atual is not None:
                push_subtotal_desp(condominio_atual, subtotal)

            params = request.GET.copy()
            params.pop("page", None)
            base_qs = params.urlencode()

            context.update({
                "despesas": despesas,
                "linhas": linhas,
                "page_obj": page_obj,
                "base_qs": base_qs,
                "total_despesas": total_despesas,
                "valor_total_despesas": valor_total_formatado,
            })

        return render(request, self.template_name, context)

class ImportarXlsxReceitasView(View):
    def post(self, request):
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            messages.error(request, "❌ Nenhum arquivo selecionado")
            return redirect("dashboard_receitas")

        if not arquivo.name.endswith((".xlsx", ".xls")):
            messages.error(request, "❌ Formato de arquivo inválido. Use XLSX ou XLS")
            return redirect("dashboard_receitas")

        try:
            workbook = openpyxl.load_workbook(arquivo)
            sheet = workbook.active

            receitas_importadas = 0
            receitas_atualizadas = 0
            erros = []

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nome_condominio = row[0]
                    unidade = row[1]
                    contato = row[2]
                    descricao_taxa = row[3]
                    cobranca = row[4]
                    valor = row[5]
                    vencimento = row[6]
                    liquidacao = row[7] if len(row) > 7 else None

                    if not nome_condominio or not unidade or valor in (None, "") or not vencimento:
                        erros.append(f"Linha {row_num}: Campos obrigatórios faltando")
                        continue

                    vencimento = to_date(vencimento)

                    if liquidacao not in (None, ""):
                        liquidacao = to_date(liquidacao)
                    else:
                        liquidacao = None

                    valor_dec = to_decimal(valor)

                    receita, created = Receitas.objects.update_or_create(
                        NomeCondominio=str(nome_condominio).strip(),
                        Unidade=str(unidade).strip(),
                        Vencimento=vencimento,
                        defaults={
                            "Contato": str(contato).strip() if contato else "",
                            "DescricaoTaxa": str(descricao_taxa).strip() if descricao_taxa else "",
                            "Cobranca": int(cobranca) if cobranca not in (None, "") else 0,
                            "Valor": valor_dec,
                            "Liquidacao": liquidacao,
                        }
                    )

                    if created:
                        receitas_importadas += 1
                    else:
                        receitas_atualizadas += 1

                except Exception as e:
                    erros.append(f"Linha {row_num}: {str(e)}")

            if receitas_importadas:
                messages.success(request, f"✅ {receitas_importadas} receita(s) importada(s) com sucesso")

            if receitas_atualizadas:
                messages.info(request, f"ℹ️ {receitas_atualizadas} receita(s) atualizada(s)")

            if erros:
                for erro in erros[:8]:
                    messages.warning(request, f"⚠️ {erro}")
                if len(erros) > 8:
                    messages.warning(request, f"⚠️ ... e mais {len(erros) - 8} erro(s)")

        except Exception as e:
            messages.error(request, f"❌ Erro ao processar arquivo: {str(e)}")

        return redirect("/dashboard/?tab=receitas")

    def get(self, request):
        return redirect("/dashboard/?tab=receitas")

class ImportarXlsxDespesasView(View):
    def post(self, request):
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            messages.error(request, "❌ Nenhum arquivo selecionado")
            return redirect("/dashboard/?tab=despesas")

        if not arquivo.name.endswith((".xlsx", ".xls")):
            messages.error(request, "❌ Formato de arquivo inválido. Use XLSX ou XLS")
            return redirect("/dashboard/?tab=despesas")

        try:
            wb = openpyxl.load_workbook(arquivo, data_only=True)

            despesas_importadas = 0
            despesas_atualizadas = 0
            erros = []

            encontrou_alguma_tabela = False

            aliases = {
                "vencimento": "vencimento",
                "liquidacao": "liquidacao",
                "liquidação": "liquidacao",
                "fornecedor": "fornecedor",
                "categoria": "categoria",
                "complemento": "complemento",
                "documento": "documento",
                "compet": "competencia",
                "compet.": "competencia",
                "competencia": "competencia",
                "forma de pagto": "formapagamento",
                "forma de pagto.": "formapagamento",
                "forma de pagamento": "formapagamento",
                "valor": "valor",
            }

            required = {"vencimento", "fornecedor", "categoria", "documento", "competencia", "valor"}

            def explode_tabs(row):
                non_empty = [c for c in row if c not in (None, "") and str(c).strip() != ""]
                if len(non_empty) == 1 and isinstance(non_empty[0], str) and "\t" in non_empty[0]:
                    parts = [p.strip() for p in non_empty[0].split("\t")]
                    return parts
                return list(row)

            def build_header_map(row):
                m = {}
                for idx, col in enumerate(row):
                    key = normalize_header(col)
                    if not key:
                        continue
                    for k, field in aliases.items():
                        nk = normalize_header(k)
                        if nk and nk in key:
                            m[field] = idx
                            break
                if not required.issubset(set(m.keys())):
                    return None
                return m

            def safe_get(row, header_map, field):
                idx = header_map.get(field)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            def is_subtotal_row(row, header_map):
                fornecedor = safe_get(row, header_map, "fornecedor")
                categoria = safe_get(row, header_map, "categoria")
                documento = safe_get(row, header_map, "documento")
                vencimento = safe_get(row, header_map, "vencimento")
                competencia = safe_get(row, header_map, "competencia")
                valor = safe_get(row, header_map, "valor")

                campos_vazios = all(x in (None, "") for x in [fornecedor, categoria, documento, vencimento, competencia])
                tem_valor = valor not in (None, "") and str(valor).strip() != ""
                return campos_vazios and tem_valor

            for sheet in wb.worksheets:
                condominio_atual = None
                header_map = None

                for row_num, raw_row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    row = explode_tabs(list(raw_row))

                    if is_row_empty(row):
                        continue

                    candidato = None
                    for c in row[:15]:
                        cand = detect_condominio(c)
                        if cand:
                            candidato = cand
                            break

                    if candidato and "vencimento" not in normalize_header(candidato):
                        condominio_atual = candidato
                        header_map = None
                        continue

                    if header_map is None:
                        hm = build_header_map(row)
                        if hm:
                            header_map = hm
                            encontrou_alguma_tabela = True
                        continue

                    if not condominio_atual:
                        erros.append(f"Aba {sheet.title} linha {row_num}: não encontrei o condomínio antes desta tabela")
                        continue

                    if is_subtotal_row(row, header_map):
                        continue

                    try:
                        vencimento = safe_get(row, header_map, "vencimento")
                        fornecedor = safe_get(row, header_map, "fornecedor")
                        categoria = safe_get(row, header_map, "categoria")
                        documento = safe_get(row, header_map, "documento")
                        competencia = safe_get(row, header_map, "competencia")
                        valor = safe_get(row, header_map, "valor")

                        if vencimento in (None, "") and fornecedor in (None, "") and valor in (None, ""):
                            continue

                        if (
                            vencimento in (None, "") or
                            fornecedor in (None, "") or
                            categoria in (None, "") or
                            documento in (None, "") or
                            competencia in (None, "") or
                            valor in (None, "")
                        ):
                            continue

                        liquidacao = safe_get(row, header_map, "liquidacao")
                        complemento = safe_get(row, header_map, "complemento")
                        formapagamento = safe_get(row, header_map, "formapagamento")

                        vencimento = to_date(vencimento)
                        liquidacao = to_date(liquidacao) if liquidacao not in (None, "") else None

                        comp_raw = competencia
                        if isinstance(comp_raw, datetime):
                            competencia_dt = comp_raw.date().replace(day=1)
                        else:
                            scomp = str(comp_raw).strip()
                            try:
                                competencia_dt = to_date(scomp).replace(day=1)
                            except Exception:
                                m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", scomp)
                                if not m:
                                    raise ValueError(f"Competência inválida: {comp_raw}")
                                mes = int(m.group(1))
                                ano = int(m.group(2))
                                competencia_dt = datetime(ano, mes, 1).date()

                        valor_dec = to_decimal(valor)

                        obj, created = Despesas.objects.update_or_create(
                            Condominio=str(condominio_atual).strip(),
                            Vencimento=vencimento,
                            Documento=str(documento).strip(),
                            Fornecedor=str(fornecedor).strip(),
                            Categoria=str(categoria).strip(),
                            defaults={
                                "Liquidacao": liquidacao,
                                "Complemento": str(complemento).strip() if complemento not in (None, "") else None,
                                "Competencia": competencia_dt,
                                "FormaPagamento": str(formapagamento).strip() if formapagamento not in (None, "") else None,
                                "Valor": valor_dec,
                            }
                        )

                        if created:
                            despesas_importadas += 1
                        else:
                            despesas_atualizadas += 1

                    except Exception as e:
                        erros.append(f"Aba {sheet.title} linha {row_num} (condomínio: {condominio_atual}): {str(e)}")

            if not encontrou_alguma_tabela:
                messages.error(request, "❌ Não encontrei o cabeçalho da tabela de despesas (Vencimento, Fornecedor, Categoria, Documento, Compet., Valor).")
                return redirect("/dashboard/?tab=despesas")

            if despesas_importadas:
                messages.success(request, f"✅ {despesas_importadas} despesa(s) importada(s) com sucesso")

            if despesas_atualizadas:
                messages.info(request, f"ℹ️ {despesas_atualizadas} despesa(s) atualizada(s)")

            if despesas_importadas == 0 and despesas_atualizadas == 0:
                messages.warning(request, "⚠️ Encontrei a tabela, mas nenhuma linha de lançamento foi detectada (sobraram só subtotais/ruídos).")

            if erros:
                for erro in erros[:10]:
                    messages.warning(request, f"⚠️ {erro}")
                if len(erros) > 10:
                    messages.warning(request, f"⚠️ ... e mais {len(erros) - 10} erro(s)")

        except Exception as e:
            messages.error(request, f"❌ Erro ao processar arquivo: {str(e)}")

        return redirect("/dashboard/?tab=despesas")

    def get(self, request):
        return redirect("/dashboard/?tab=despesas")
